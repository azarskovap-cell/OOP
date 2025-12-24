import openpyxl  # для Excel
from openpyxl.styles import Font  # для жирного текста


class ExcelReport:
    def __init__(self, journal):
        self.journal = journal  # берем журнал

    def generate(self, start_date, end_date, filename='report.xlsx'):
        wb = openpyxl.Workbook()  # новая книга Excel
        ws = wb.active  # активный лист
        ws.title = "Отчёт по нагрузке"  # название листа

        # шапка
        ws['A1'] = f"Преподаватель: {self.journal.teacher.get_name()}"
        ws['A2'] = f"Период: {start_date} - {end_date}"

        # заголовки таблицы
        headers = ['Дата', 'Группа', 'Дисциплина', 'Тема занятия', 'Часы']
        for col, header in enumerate(headers, 1):  # колонки 1,2,3,4,5
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)  # жирный текст

        # данные
        lessons = self.journal.get_lessons_by_date(start_date, end_date)
        row = 5  # начинаем с 5 строки
        for lesson in lessons:
            ws.cell(row=row, column=1, value=lesson.date)
            ws.cell(row=row, column=2, value=lesson.group.get_name())
            ws.cell(row=row, column=3, value=lesson.discipline.get_name())
            ws.cell(row=row, column=4, value=lesson.topic)
            ws.cell(row=row, column=5, value=lesson.hours)
            row += 1  # следующая строка

        # итоги
        total_row = row + 1
        ws.cell(row=total_row, column=4, value="Итого часов:").font = Font(bold=True)
        total_hours = self.journal.calculate_total_hours(start_date, end_date)
        ws.cell(row=total_row, column=5, value=total_hours).font = Font(bold=True)

        # сохраняем на Рабочий стол
        import os
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        full_path = os.path.join(desktop, filename)
        wb.save(full_path)  # сохраняем на Рабочий стол
        return full_path  # возвращаем полный путь